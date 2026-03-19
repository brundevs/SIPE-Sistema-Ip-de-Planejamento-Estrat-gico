"""
SIPE | Sistema Ipê de Planejamento Estratégico - Rotas da API Web (Flask)
Endpoints para Efetivo, RDO e Clima.
"""
import json
import os
from pathlib import Path
from datetime import datetime, timezone

from flask import Blueprint, request, jsonify, current_app, send_from_directory
from sqlalchemy import func
from werkzeug.utils import secure_filename
import openpyxl

from app.database.models import Colaborador, Vinculo, ProcessamentoRDO, LogAtividade, HistoricoTerceiro, Projeto, Tarefa, Equipamento, Veiculo, HistoricoLiberacao, PermissaoTrabalho
from app.database.session import SessionLocal
from app.core.pdf_extractor import extrair_texto_pdf, extrair_data_documento, extrair_colaboradores_pte, extrair_cpfs, extrair_horarios_pte, extrair_nomes_do_rdo, extrair_permissoes_trabalho
from app.core.fuzzy_engine import processar_lista_nomes, buscar_melhor_match
from app.services.clima_service import obter_clima
from app.services.log_service import registrar_log

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from config import (
    UPLOADS_DIR, PDFS_DIR, ALLOWED_EXTENSIONS_PDF, ALLOWED_EXTENSIONS_EXCEL,
    VIP_NAMES
)

api = Blueprint("api", __name__, url_prefix="/api")


# ─────────────────────────────────────────────
# EFETIVO (Gestão de Colaboradores)
# ─────────────────────────────────────────────

import re as _re

def _limpar_cpf(valor: str) -> str:
    """Remove tudo que não for dígito do CPF. Retorna string com 11 dígitos ou menos."""
    return _re.sub(r'\D', '', str(valor or ''))

def _normalizar_nome(valor: str) -> str:
    """Aplica Title Case e remove espaços extras."""
    if not valor:
        return ''
    return _re.sub(r'\s+', ' ', str(valor).strip()).title()

def _normalizar_categoria(valor: str):
    """Aceita MOD, MOI (case insensitive); retorna em maiúsculas ou None."""
    if not valor:
        return None
    v = _re.sub(r'\s+', '', str(valor)).upper()
    return v if v in ('MOD', 'MOI') else None


@api.route("/efetivo/upload-excel", methods=["POST"])
def upload_excel():
    """Upload e importação de planilha Excel com dados de colaboradores."""
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Nome de arquivo vazio"}), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS_EXCEL:
        return jsonify({"error": f"Extensão inválida. Use: {ALLOWED_EXTENSIONS_EXCEL}"}), 400

    db = SessionLocal()
    try:
        filename = secure_filename(file.filename)
        filepath = UPLOADS_DIR / filename
        file.save(str(filepath))

        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        ws = wb.active

        # ── Detectar colunas (cabeçalhos flexíveis) ────────────
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                h = str(cell.value).strip().upper()
                if "NOME" in h or "FUNCIONÁRIO" in h or "COLABORADOR" in h:
                    headers["nome"] = col_idx
                elif "CPF" in h:
                    headers["cpf"] = col_idx
                elif "MATR" in h or "REG" in h or h in ("MAT", "RE", "CHAPA"):
                    headers["matricula"] = col_idx
                elif "CARGO" in h or "FUNÇÃO" in h or "FUNCAO" in h:
                    headers["cargo"] = col_idx
                elif "SETOR" in h or "DEPARTAMENTO" in h or "DEPTO" in h:
                    headers["setor"] = col_idx
                elif "CATEGORIA" in h or h in ("MOD", "MOI", "TIPO"):
                    headers["categoria"] = col_idx

        if "nome" not in headers:
            return jsonify({"error": "Coluna de NOME não encontrada. Verifique os cabeçalhos."}), 400

        importados = 0
        atualizados = 0
        erros = 0
        avisos = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                nome_cell = row[headers["nome"] - 1].value
                if not nome_cell or not str(nome_cell).strip():
                    continue

                # ── Padronização ───────────────────────────────
                nome = _normalizar_nome(str(nome_cell))

                # CPF: apenas dígitos, validar 11 chars
                cpf = None
                if "cpf" in headers and row[headers["cpf"] - 1].value:
                    cpf_raw = str(row[headers["cpf"] - 1].value).strip()
                    cpf_limpo = _limpar_cpf(cpf_raw)
                    if len(cpf_limpo) == 11:
                        cpf = cpf_limpo
                    else:
                        avisos.append(
                            f"Linha {row_idx}: CPF '{cpf_raw}' ignorado — deve ter 11 dígitos."
                        )

                matricula = None
                if "matricula" in headers and row[headers["matricula"] - 1].value:
                    matricula = str(row[headers["matricula"] - 1].value).strip()

                # Cargo: Title Case
                cargo = None
                if "cargo" in headers and row[headers["cargo"] - 1].value:
                    cargo = str(row[headers["cargo"] - 1].value).strip().title()

                setor = None
                if "setor" in headers and row[headers["setor"] - 1].value:
                    setor = str(row[headers["setor"] - 1].value).strip()

                # Categoria: MOD ou MOI
                categoria = None
                if "categoria" in headers and row[headers["categoria"] - 1].value:
                    cat_raw = str(row[headers["categoria"] - 1].value).strip()
                    categoria = _normalizar_categoria(cat_raw)
                    if not categoria:
                        avisos.append(
                            f"Linha {row_idx}: Categoria '{cat_raw}' ignorada — use MOD ou MOI."
                        )

                # ── Persistência ───────────────────────────────
                existente = None
                if cpf:
                    existente = db.query(Colaborador).filter(Colaborador.cpf == cpf).first()
                if not existente and matricula:
                    existente = db.query(Colaborador).filter(Colaborador.matricula == matricula).first()
                if not existente:
                    existente = db.query(Colaborador).filter(
                        func.lower(Colaborador.nome) == nome.lower()
                    ).first()

                if existente:
                    existente.nome = nome
                    if cpf:       existente.cpf = cpf
                    if matricula: existente.matricula = matricula
                    if cargo:     existente.cargo = cargo
                    if setor:     existente.setor = setor
                    if categoria: existente.categoria = categoria
                    existente.data_atualizacao = datetime.now(timezone.utc)
                    atualizados += 1
                else:
                    novo = Colaborador(
                        nome=nome, cpf=cpf, matricula=matricula,
                        cargo=cargo, setor=setor, categoria=categoria
                    )
                    db.add(novo)
                    importados += 1

            except Exception:
                erros += 1
                continue

        db.commit()

        registrar_log(db, "success", "efetivo",
                      f"Excel importado: {importados} novos, {atualizados} atualizados, {erros} erros",
                      f"Arquivo: {filename}")

        os.remove(str(filepath))

        return jsonify({
            "success": True,
            "importados": importados,
            "atualizados": atualizados,
            "erros": erros,
            "avisos": avisos,
            "total_base": db.query(Colaborador).count()
        })

    except Exception as e:
        db.rollback()
        return jsonify({"error": f"Erro ao processar Excel: {str(e)}"}), 500
    finally:
        db.close()



@api.route("/efetivo/modelo-padrao", methods=["GET"])
def baixar_modelo_padrao():
    """
    Gera e devolve um arquivo .xlsx de template para importação de
    colaboradores, com formatação, exemplos e validação de Categoria.
    """
    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── Aba Colaboradores ──────────────────────────────────────
    ws = wb.active
    ws.title = "Colaboradores"

    verde   = "1A6B3C"   # verde Ipê
    verde_l = "EAF4EE"   # verde claro
    branco  = "FFFFFF"

    thin   = Side(style="thin", color="CCCCCC")
    borda  = Border(left=thin, right=thin, top=thin, bottom=thin)
    centro = Alignment(horizontal="center", vertical="center")
    topo   = Alignment(wrap_text=True, vertical="top")

    cabecalhos = [
        ("Nome Completo", 42),
        ("CPF",           22),
        ("Cargo",         30),
        ("Categoria",     16),
    ]

    for col, (titulo, larg) in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font      = Font(bold=True, color=branco, size=11)
        c.fill      = PatternFill("solid", fgColor=verde)
        c.alignment = centro
        c.border    = borda
        ws.column_dimensions[get_column_letter(col)].width = larg
    ws.row_dimensions[1].height = 24

    exemplos = [
        ("Ana Paula Bastos",  "000.000.000-00",  "Auxiliar Administrativo", "MOI"),
        ("Bruno Silva Lima",  "11122233344",     "Técnico Mecânico",        "MOD"),
        ("Carla Souza Costa", "55566677788",     "Encarregado",             "MOD"),
    ]
    fill_ex = PatternFill("solid", fgColor=verde_l)
    for r, linha in enumerate(exemplos, 2):
        for c, val in enumerate(linha, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill      = fill_ex
            cell.alignment = topo
            cell.border    = borda
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"

    # Validação de lista para Categoria (coluna D)
    dv = DataValidation(type="list", formula1='"MOD,MOI"', allow_blank=True)
    dv.sqref = "D2:D10000"
    ws.add_data_validation(dv)

    # ── Aba Instruções ─────────────────────────────────────────
    wi = wb.create_sheet("Instruções")
    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 65

    cabec_f = Font(bold=True, color=branco, size=11)
    cabec_p = PatternFill("solid", fgColor=verde)
    titulo_f = Font(bold=True, size=13, color=verde)
    normal_f = Font(size=11)
    wrap_a   = Alignment(wrap_text=True, vertical="top")

    wi["A1"] = "MODELO PADRÃO DE IMPORTAÇÃO — SIPE"
    wi["A1"].font = titulo_f
    wi.merge_cells("A1:B1")
    wi["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wi.row_dimensions[1].height = 28

    instrucoes = [
        ("Campo",          "Descrição / Regras"),
        ("Nome Completo",  "Nome sem abreviações. O sistema normaliza para Title Case: 'BRUNO BASTOS' vira 'Bruno Bastos'."),
        ("CPF",            "Aceita com ou sem máscara. O sistema remove pontos/traços e valida 11 dígitos."),
        ("Cargo",          "Nomenclatura oficial. O sistema normalizará as maiúsculas automaticamente."),
        ("Categoria",      "\"MOD\" (Mão de Obra Direta) ou \"MOI\" (Mão de Obra Indireta). Outros valores são descartados com aviso."),
    ]
    for r, (campo, desc) in enumerate(instrucoes, 3):
        ca = wi.cell(row=r, column=1, value=campo)
        cb = wi.cell(row=r, column=2, value=desc)
        if r == 3:
            ca.font = cabec_f; ca.fill = cabec_p
            cb.font = cabec_f; cb.fill = cabec_p
        else:
            ca.font = Font(bold=True, size=11)
            cb.font = normal_f
        ca.alignment = wrap_a
        cb.alignment = wrap_a
        wi.row_dimensions[r].height = 42

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="modelo_importacao_colaboradores.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@api.route("/efetivo/colaboradores", methods=["GET"])
def listar_colaboradores():
    """Lista todos os colaboradores cadastrados."""
    db = SessionLocal()
    try:
        busca = request.args.get("busca", "").strip()
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 50))

        query = db.query(Colaborador).filter(Colaborador.ativo == True)

        if busca:
            query = query.filter(
                Colaborador.nome.ilike(f"%{busca}%") |
                Colaborador.cpf.ilike(f"%{busca}%") |
                Colaborador.matricula.ilike(f"%{busca}%")
            )

        total = query.count()
        colaboradores = query.order_by(Colaborador.nome).offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            "colaboradores": [c.to_dict() for c in colaboradores],
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": (total + per_page - 1) // per_page
        })
    finally:
        db.close()


@api.route("/efetivo/colaboradores/<int:colab_id>", methods=["DELETE"])
def remover_colaborador(colab_id):
    """Remove (desativa) um colaborador."""
    db = SessionLocal()
    try:
        colab = db.query(Colaborador).get(colab_id)
        if not colab:
            return jsonify({"error": "Colaborador não encontrado"}), 404
        colab.ativo = False
        db.commit()
        registrar_log(db, "info", "efetivo", f"Colaborador desativado: {colab.nome}")
        return jsonify({"success": True, "mensagem": f"Colaborador {colab.nome} desativado"})
    finally:
        db.close()


@api.route("/efetivo/estatisticas", methods=["GET"])
def estatisticas_efetivo():
    """Retorna estatísticas do efetivo."""
    db = SessionLocal()
    try:
        total = db.query(Colaborador).filter(Colaborador.ativo == True).count()
        total_vinculos = db.query(Vinculo).count()
        vinculos_confirmados = db.query(Vinculo).filter(Vinculo.confirmado == True).count()

        return jsonify({
            "total_colaboradores": total,
            "total_vinculos": total_vinculos,
            "vinculos_confirmados": vinculos_confirmados,
        })
    finally:
        db.close()


@api.route("/efetivo/adicionar", methods=["POST"])
def adicionar_colaborador():
    """Adiciona um colaborador manualmente (sem planilha)."""
    db = SessionLocal()
    try:
        dados = request.get_json(force=True) or {}
        nome = _normalizar_nome(dados.get("nome", ""))
        if not nome:
            return jsonify({"error": "Nome é obrigatório"}), 400

        cpf_raw = dados.get("cpf", "")
        cpf = _limpar_cpf(cpf_raw) if cpf_raw else None
        if cpf and len(cpf) != 11:
            return jsonify({"error": "CPF deve ter 11 dígitos"}), 400

        empresa = str(dados.get("empresa", "") or "").strip() or None
        cargo_raw = str(dados.get("cargo", "") or "").strip()
        cargo = cargo_raw.title() if cargo_raw else None

        # Verifica duplicidade por CPF ou nome
        existente = None
        if cpf:
            existente = db.query(Colaborador).filter(Colaborador.cpf == cpf).first()
        if not existente:
            existente = db.query(Colaborador).filter(
                func.lower(Colaborador.nome) == nome.lower()
            ).first()

        if existente:
            existente.nome = nome
            if cpf: existente.cpf = cpf
            if empresa: existente.empresa = empresa
            if cargo: existente.cargo = cargo
            existente.data_atualizacao = datetime.now(timezone.utc)
            db.commit()
            registrar_log(db, "info", "efetivo", f"Colaborador atualizado manualmente: {nome}")
            return jsonify({"success": True, "acao": "atualizado", "id": existente.id})
        else:
            novo = Colaborador(nome=nome, cpf=cpf, empresa=empresa, cargo=cargo)
            db.add(novo)
            db.commit()
            registrar_log(db, "success", "efetivo", f"Colaborador adicionado manualmente: {nome}")
            return jsonify({"success": True, "acao": "criado", "id": novo.id})

    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/efetivo/exportar", methods=["GET"])
def exportar_efetivo_xlsx():
    """Exporta todos os colaboradores ativos como arquivo .xlsx."""
    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = SessionLocal()
    try:
        colaboradores = db.query(Colaborador).filter(Colaborador.ativo == True).order_by(Colaborador.nome).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Efetivo"

        verde = "1A6B3C"
        thin = Side(style="thin", color="CCCCCC")
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Nome", "CPF", "Empresa", "Cargo", "Setor", "Categoria", "Data Cadastro"]
        widths = [40, 16, 30, 30, 25, 12, 22]

        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor=verde)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 24

        def fmt_cpf(c):
            if c and len(c) == 11:
                return f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}"
            return c or ""

        for ri, col in enumerate(colaboradores, 2):
            data_cad = col.data_importacao.strftime("%d/%m/%Y") if col.data_importacao else ""
            row_data = [col.nome, fmt_cpf(col.cpf), col.empresa or "", col.cargo or "",
                        col.setor or "", col.categoria or "", data_cad]
            fill_color = "EAF4EE" if ri % 2 == 0 else "FFFFFF"
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = borda
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[ri].height = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        hoje = datetime.now().strftime("%Y%m%d")
        return send_file(buf, as_attachment=True,
                         download_name=f"efetivo_{hoje}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        db.close()


# ─────────────────────────────────────────────
# LIBERAÇÃO DE ACESSOS
# ─────────────────────────────────────────────

@api.route("/acesso/buscar", methods=["GET"])
def buscar_acesso():
    """Busca pessoas em colaboradores e histórico de terceiros para auto-complete."""
    db = SessionLocal()
    try:
        q = request.args.get("q", "").strip()
        if len(q) < 2:
            return jsonify({"resultados": []})

        resultados = []

        # Busca em colaboradores
        colabs = db.query(Colaborador).filter(
            Colaborador.ativo == True,
            (Colaborador.nome.ilike(f"%{q}%") | Colaborador.cpf.ilike(f"%{q}%"))
        ).order_by(Colaborador.nome).limit(8).all()
        for c in colabs:
            resultados.append({
                "nome": c.nome, "cpf": c.cpf or "", "empresa": c.empresa or "",
                "cargo": c.cargo or "", "placa": "", "local": "", "motivo": "",
                "origem": "efetivo"
            })

        # Busca em historico_terceiros
        terceiros = db.query(HistoricoTerceiro).filter(
            HistoricoTerceiro.nome.ilike(f"%{q}%") | HistoricoTerceiro.cpf.ilike(f"%{q}%")
        ).order_by(HistoricoTerceiro.data_cadastro.desc()).limit(8).all()
        for t in terceiros:
            # Evitar duplicatas por CPF
            if t.cpf and any(r["cpf"] == t.cpf for r in resultados):
                continue
            resultados.append({
                "nome": t.nome, "cpf": t.cpf or "", "empresa": t.empresa or "",
                "cargo": "", "placa": t.placa or "", "local": t.local or "",
                "motivo": t.motivo or "", "origem": "terceiro"
            })

        return jsonify({"resultados": resultados[:10]})
    finally:
        db.close()


@api.route("/acesso/liberar", methods=["POST"])
def liberar_acesso():
    """Registra liberação de acesso e auto-salva novo terceiro se não existir."""
    db = SessionLocal()
    try:
        dados = request.get_json(force=True) or {}
        nome = _normalizar_nome(dados.get("motorista", ""))
        cpf = _limpar_cpf(dados.get("cpf", "")) or None
        empresa = str(dados.get("empresa", "") or "").strip() or None
        placa = str(dados.get("placa", "") or "").strip().upper() or None

        if not nome:
            return jsonify({"error": "Nome do motorista é obrigatório"}), 400

        local_acesso = str(dados.get("local", "") or "").strip() or None
        motivo_acesso = str(dados.get("motivo", "") or "").strip() or None
        texto_gerado = str(dados.get("texto_gerado", "") or "").strip()
        gerado_por = str(dados.get("gerado_por", "") or "").strip() or None
        periodo_str = str(dados.get("periodo", "") or "").strip() or None
        data_acesso_str = str(dados.get("data_acesso", "") or "").strip() or None

        # Só salva em historico_terceiros se NÃO for do efetivo
        is_efetivo = False
        if cpf:
            is_efetivo = db.query(Colaborador).filter(
                Colaborador.cpf == cpf, Colaborador.ativo == True
            ).first() is not None
        if not is_efetivo:
            is_efetivo = db.query(Colaborador).filter(
                func.lower(Colaborador.nome) == nome.lower(), Colaborador.ativo == True
            ).first() is not None

        if not is_efetivo:
            terceiro = None
            if cpf:
                terceiro = db.query(HistoricoTerceiro).filter(HistoricoTerceiro.cpf == cpf).first()
            if not terceiro:
                terceiro = db.query(HistoricoTerceiro).filter(
                    func.lower(HistoricoTerceiro.nome) == nome.lower()
                ).first()

            if terceiro:
                if placa: terceiro.placa = placa
                if empresa: terceiro.empresa = empresa
                if local_acesso: terceiro.local = local_acesso
                if motivo_acesso: terceiro.motivo = motivo_acesso
            else:
                terceiro = HistoricoTerceiro(
                    nome=nome, cpf=cpf, placa=placa, empresa=empresa,
                    local=local_acesso, motivo=motivo_acesso
                )
                db.add(terceiro)
            db.commit()

        registrar_log(db, "success", "acesso",
                      f"Acesso liberado: {nome}",
                      f"Local: {dados.get('local','')} | Placa: {placa}")

        if texto_gerado:
            hist_lib = HistoricoLiberacao(
                motorista=nome,
                cpf=cpf,
                empresa=empresa,
                placa=placa,
                local=local_acesso,
                motivo=motivo_acesso,
                periodo=periodo_str,
                data_acesso=data_acesso_str,
                gerado_por=gerado_por,
                texto_gerado=texto_gerado,
            )
            db.add(hist_lib)
            db.commit()

        return jsonify({"success": True})

    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/acesso/historico-liberacoes", methods=["GET"])
def historico_liberacoes():
    """Lista histórico de liberações geradas."""
    db = SessionLocal()
    try:
        items = db.query(HistoricoLiberacao).order_by(
            HistoricoLiberacao.data_geracao.desc()
        ).limit(100).all()
        return jsonify({"liberacoes": [i.to_dict() for i in items]})
    finally:
        db.close()


@api.route("/acesso/historico-liberacoes/<int:lid>", methods=["DELETE"])
def deletar_liberacao(lid):
    db = SessionLocal()
    try:
        item = db.query(HistoricoLiberacao).get(lid)
        if not item:
            return jsonify({"error": "Não encontrado"}), 404
        db.delete(item)
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


# ─────────────────────────────────────────────
# PROCESSAMENTO RDO
# ─────────────────────────────────────────────

@api.route("/rdo/processar", methods=["POST"])
def processar_rdo():
    """Upload e processamento de PDF do RDO."""
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo PDF enviado"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Nome de arquivo vazio"}), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS_PDF:
        return jsonify({"error": f"Extensão inválida. Use: {ALLOWED_EXTENSIONS_PDF}"}), 400

    db = SessionLocal()
    try:
        # Salvar PDF
        filename = secure_filename(file.filename)
        filepath = UPLOADS_DIR / filename
        file.save(str(filepath))

        # Criar registro de processamento
        proc = ProcessamentoRDO(
            nome_arquivo=filename,
            status="processando"
        )
        db.add(proc)
        db.commit()

        # Extrair texto do PDF
        texto = extrair_texto_pdf(str(filepath))
        nomes_extraidos = extrair_nomes_do_rdo(texto)

        # Buscar colaboradores na base
        colaboradores = db.query(Colaborador).filter(Colaborador.ativo == True).all()
        lista_colabs = [c.to_dict() for c in colaboradores]

        # Processar com fuzzy matching
        nomes_pdf = [n["nome_limpo"] for n in nomes_extraidos]
        resultados = processar_lista_nomes(nomes_pdf, lista_colabs, VIP_NAMES)

        # Salvar vínculos automáticos
        for match in resultados["automaticos"]:
            colab = db.query(Colaborador).filter(
                func.lower(Colaborador.nome) == match["match"].lower()
            ).first()
            if colab:
                vinculo_existente = db.query(Vinculo).filter(
                    Vinculo.nome_pdf == match["nome_pdf"],
                    Vinculo.colaborador_id == colab.id
                ).first()
                if not vinculo_existente:
                    vinculo = Vinculo(
                        nome_pdf=match["nome_pdf"],
                        colaborador_id=colab.id,
                        score_similaridade=match["score"],
                        confirmado=True
                    )
                    db.add(vinculo)

        # Atualizar registro de processamento
        proc.total_nomes_extraidos = len(nomes_extraidos)
        proc.total_matches_auto = resultados["estatisticas"]["automaticos"]
        proc.total_matches_revisao = resultados["estatisticas"]["revisao"]
        proc.total_sem_match = resultados["estatisticas"]["sem_match"]
        proc.status = "concluido"
        proc.resultado_json = json.dumps(resultados, ensure_ascii=False)

        db.commit()

        registrar_log(db, "success", "rdo",
                      f"RDO processado: {filename}",
                      f"Extraídos: {len(nomes_extraidos)}, Auto: {resultados['estatisticas']['automaticos']}, Revisão: {resultados['estatisticas']['revisao']}")

        # Limpar PDF
        os.remove(str(filepath))

        return jsonify({
            "success": True,
            "processamento_id": proc.id,
            "resultado": resultados,
            "nomes_extraidos": nomes_extraidos
        })

    except Exception as e:
        db.rollback()
        if 'proc' in locals():
            proc.status = "erro"
            proc.erro_mensagem = str(e)
            db.commit()
        registrar_log(db, "error", "rdo", f"Erro ao processar RDO: {str(e)}")
        return jsonify({"error": f"Erro ao processar RDO: {str(e)}"}), 500
    finally:
        db.close()


@api.route("/rdo/confirmar-vinculo", methods=["POST"])
def confirmar_vinculo():
    """Confirma ou rejeita um vínculo sugerido."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dados inválidos"}), 400

    nome_pdf = data.get("nome_pdf")
    colaborador_id = data.get("colaborador_id")
    confirmar = data.get("confirmar", True)

    if not nome_pdf or not colaborador_id:
        return jsonify({"error": "nome_pdf e colaborador_id são obrigatórios"}), 400

    db = SessionLocal()
    try:
        if confirmar:
            vinculo = Vinculo(
                nome_pdf=nome_pdf,
                colaborador_id=colaborador_id,
                score_similaridade=data.get("score", 0),
                confirmado=True
            )
            db.add(vinculo)
            db.commit()
            registrar_log(db, "info", "rdo", f"Vínculo confirmado: '{nome_pdf}' → ID {colaborador_id}")
            return jsonify({"success": True, "mensagem": "Vínculo confirmado"})
        else:
            registrar_log(db, "info", "rdo", f"Vínculo rejeitado: '{nome_pdf}' → ID {colaborador_id}")
            return jsonify({"success": True, "mensagem": "Vínculo rejeitado"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/rdo/historico", methods=["GET"])
def historico_processamentos():
    """Lista histórico de processamentos. Com busca_nome: busca pessoa dentro dos JSONs."""
    busca = request.args.get("busca_nome", "").strip().lower()
    db = SessionLocal()
    try:
        procs = db.query(ProcessamentoRDO).order_by(ProcessamentoRDO.data_processamento.desc()).all()

        if busca:
            # Modo busca de pessoa: retorna quais arquivos contêm essa pessoa
            matches = []
            for p in procs:
                json_str = p.resultado_json or ""
                if busca not in json_str.lower():
                    continue
                # Parse para extrair nome exato e horários
                try:
                    resultado = json.loads(json_str)
                except Exception:
                    resultado = {}
                pessoas_encontradas = []
                if isinstance(resultado, dict):
                    for key, colabs in resultado.items():
                        if isinstance(colabs, list):
                            for c in colabs:
                                nome_c = (c.get('nome') or '').lower()
                                if busca in nome_c:
                                    pessoas_encontradas.append({
                                        "nome": c.get('nome', ''),
                                        "categoria": c.get('categoria', ''),
                                        "cargo": c.get('cargo', ''),
                                    })
                elif isinstance(resultado, list):
                    for entry in resultado:
                        for c in (entry.get('colaboradores') or []):
                            nome_c = (c.get('nome') or '').lower()
                            if busca in nome_c:
                                pessoas_encontradas.append({
                                    "nome": c.get('nome', ''),
                                    "categoria": c.get('categoria', ''),
                                    "cargo": c.get('cargo', ''),
                                })
                if pessoas_encontradas:
                    d = p.to_dict()
                    d['pessoas_encontradas'] = pessoas_encontradas
                    matches.append(d)
            return jsonify({"processamentos": matches, "modo_busca_pessoa": True, "busca": busca})

        # Modo lista normal
        result = []
        for p in procs:
            result.append(p.to_dict())
            if len(result) >= 50:
                break
        return jsonify({"processamentos": result})
    finally:
        db.close()


@api.route("/rdo/historico/<int:proc_id>", methods=["DELETE"])
def deletar_processamento(proc_id):
    """Remove um registro do histórico de processamentos."""
    db = SessionLocal()
    try:
        proc = db.query(ProcessamentoRDO).filter(ProcessamentoRDO.id == proc_id).first()
        if not proc:
            return jsonify({"error": "Processamento não encontrado"}), 404
        nome = proc.nome_arquivo
        db.delete(proc)
        db.commit()
        registrar_log(db, "info", "rdo", f"Histórico removido: {nome}")
        return jsonify({"success": True, "mensagem": f"Registro '{nome}' removido."})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/efetivo/colaboradores/<int:colab_id>/categoria", methods=["PUT"])
def atualizar_categoria(colab_id):
    """Atualiza a categoria (MOD/MOI) de um colaborador."""
    data = request.get_json()
    nova_categoria = data.get("categoria", "").strip().upper()
    if nova_categoria not in ["MOD", "MOI"]:
        return jsonify({"error": "Categoria inválida. Use MOD ou MOI."}), 400

    db = SessionLocal()
    try:
        colab = db.query(Colaborador).get(colab_id)
        if not colab or not colab.ativo:
            return jsonify({"error": "Colaborador não encontrado"}), 404
        
        colab.categoria = nova_categoria
        db.commit()
        registrar_log(db, "info", "efetivo", f"Categoria atualizada para {nova_categoria}: {colab.nome}")
        return jsonify({"success": True, "categoria": nova_categoria})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


# ─────────────────────────────────────────────
# PTE / CESLA (Leitura de Presença MOD)
# ─────────────────────────────────────────────

@api.route("/pte/processar", methods=["POST"])
def processar_pte():
    """
    Recebe um ou mais PDFs (campo 'files[]') do PTE/Cesla,
    extrai colaboradores MOD comparando com a Base,
    e injeta automaticamente TODOS os colaboradores MOI ativos.
    """
    arquivos = request.files.getlist("files[]")
    if not arquivos:
        arquivo_single = request.files.get("file")
        if arquivo_single:
            arquivos = [arquivo_single]
        else:
            return jsonify({"error": "Nenhum arquivo PDF enviado"}), 400

    resultados = []
    erros = []

    db = SessionLocal()
    try:
        # Pré-carregar a base de colaboradores ativos
        colaboradores_db = db.query(Colaborador).filter(Colaborador.ativo == True).all()
        base_dicts = [c.to_dict() for c in colaboradores_db]
        nomes_base = [c['nome'] for c in base_dicts]
        
        # Mapeia por CPF e também por Matrícula (caso o CPF tenha sido salvo na coluna de matrícula)
        mapa_cpf = {}
        for c in base_dicts:
            if c.get('cpf'):
                mapa_cpf[_limpar_cpf(c['cpf'])] = c
            if c.get('matricula'):
                mat_limpa = _limpar_cpf(c['matricula'])
                if len(mat_limpa) == 11:
                    mapa_cpf[mat_limpa] = c
                    
        mapa_nome = {c['nome']: c for c in base_dicts}
        mapa_mat = {str(c['matricula']).strip(): c for c in base_dicts if c.get('matricula')}

        # Coletar a lista de todos os MOI
        mois_ativos = [c for c in base_dicts if c.get('categoria') == 'MOI']

        for file in arquivos:
            if not file or not file.filename:
                continue

            ext = Path(file.filename).suffix.lower()
            if ext not in ALLOWED_EXTENSIONS_PDF:
                erros.append({"arquivo": file.filename, "erro": "Extensão inválida (apenas .pdf)"})
                continue

            try:
                filename = secure_filename(file.filename)
                filepath = UPLOADS_DIR / filename
                file.save(str(filepath))

                texto = extrair_texto_pdf(str(filepath))
                data_doc = extrair_data_documento(texto)
                hr_inicio, hr_fim = extrair_horarios_pte(texto)
                cands_heuristica = extrair_colaboradores_pte(texto)

                colabs_encontrados = {}

                # 1. Match candidatos da heurística contra a Base
                for cand in cands_heuristica:
                    cpf_cand = _limpar_cpf(cand.get('cpf', ''))
                    if cpf_cand and cpf_cand in mapa_cpf:
                        c = mapa_cpf[cpf_cand]
                        colabs_encontrados[c['id']] = {
                            "id": c['id'],
                            "nome": c['nome'],
                            "cpf": c.get('cpf') or c.get('matricula') or '',
                            "matricula": c.get('matricula') or c.get('cpf') or '',
                            "cargo": c.get('cargo', cand.get('cargo', '')),
                            "categoria": c.get('categoria', 'MOD')
                        }
                        continue

                    mat_cand = cand.get('matricula', '').strip()
                    if mat_cand and mat_cand in mapa_mat:
                        c = mapa_mat[mat_cand]
                        colabs_encontrados[c['id']] = {
                            "id": c['id'],
                            "nome": c['nome'],
                            "cpf": c.get('cpf') or c.get('matricula') or '',
                            "matricula": c.get('matricula') or c.get('cpf') or '',
                            "cargo": c.get('cargo', cand.get('cargo', '')),
                            "categoria": c.get('categoria', 'MOD')
                        }
                        continue
                    
                    nm_cand = cand.get('nome', '')
                    if nm_cand:
                        match = buscar_melhor_match(nm_cand, nomes_base, limit=1)
                        if match['score'] >= 65:
                            c = mapa_nome[match['melhor_match']]
                            colabs_encontrados[c['id']] = {
                                "id": c['id'],
                                "nome": c['nome'],
                                "cpf": c.get('cpf') or c.get('matricula') or '',
                                "matricula": c.get('matricula') or c.get('cpf') or '',
                                "cargo": c.get('cargo', cand.get('cargo', '')),
                                "categoria": c.get('categoria', 'MOD')
                            }

                # 2. Match por CPFs extraídos do texto inteiro
                cpfs_inteiros = extrair_cpfs(texto)
                for cpf_raw in cpfs_inteiros:
                    cpf_limpo = _limpar_cpf(cpf_raw)
                    if cpf_limpo in mapa_cpf:
                        c = mapa_cpf[cpf_limpo]
                        colabs_encontrados[c['id']] = {
                            "id": c['id'],
                            "nome": c['nome'],
                            "cpf": c.get('cpf') or c.get('matricula') or '',
                            "matricula": c.get('matricula') or c.get('cpf') or '',
                            "cargo": c.get('cargo', ''),
                            "categoria": c.get('categoria', 'MOD')
                        }

                # 3. Match por Nome Exato da Base
                texto_upper = texto.upper()
                for c in base_dicts:
                    if c['id'] not in colabs_encontrados:
                        if len(c['nome']) > 5 and c['nome'].upper() in texto_upper:
                             colabs_encontrados[c['id']] = {
                                "id": c['id'],
                                "nome": c['nome'],
                                "cpf": c.get('cpf') or c.get('matricula') or '',
                                "matricula": c.get('matricula') or c.get('cpf') or '',
                                "cargo": c.get('cargo', ''),
                                "categoria": c.get('categoria', 'MOD')
                            }

                # 4. Injetar todos os MOI ativos
                for moi in mois_ativos:
                    if moi['id'] not in colabs_encontrados:
                        colabs_encontrados[moi['id']] = {
                            "id": moi['id'],
                            "nome": moi['nome'],
                            "cpf": moi.get('cpf') or moi.get('matricula') or '',
                            "matricula": moi.get('matricula') or moi.get('cpf') or '',
                            "cargo": moi.get('cargo', ''),
                            "categoria": "MOI"
                        }

                lista_final = list(colabs_encontrados.values())
                lista_final.sort(key=lambda x: x['nome'])

                # Mover PDF para armazenamento permanente
                import shutil
                pdf_filename = None
                try:
                    import time as _time
                    safe_name = f"{int(_time.time()*1000)}_{filename}"
                    dest = PDFS_DIR / safe_name
                    shutil.move(str(filepath), str(dest))
                    pdf_filename = safe_name
                except Exception:
                    try:
                        os.remove(str(filepath))
                    except Exception:
                        pass

                # 5. Extrair Permissões de Trabalho do PDF
                permissoes_pt = extrair_permissoes_trabalho(texto)

                resultados.append({
                    "arquivo": file.filename,
                    "data": data_doc,
                    "inicio": hr_inicio,
                    "fim": hr_fim,
                    "total": len(lista_final),
                    "colaboradores": lista_final,
                    "pdf_filename": pdf_filename,
                    "permissoes": permissoes_pt,
                })

            except Exception as e:
                erros.append({"arquivo": file.filename, "erro": str(e)})
                try:
                    os.remove(str(filepath))
                except Exception:
                    pass

        if not resultados and erros:
            return jsonify({"error": "Todos os arquivos falharam", "detalhes": erros}), 400

        total_colabs = sum(r["total"] for r in resultados)
        registrar_log(
            db, "success", "pte",
            f"PTE processado: {len(resultados)} arquivo(s), {total_colabs} colaborador(es) MOD+MOI",
            f"Arquivos: {[r['arquivo'] for r in resultados]}"
        )

        return jsonify({
            "success": True,
            "processados": len(resultados),
            "resultados": resultados,
            "erros": erros,
        })
    finally:
        db.close()


@api.route("/pte/confirmar", methods=["POST"])
def confirmar_pte():
    """
    Recebe os resultados confirmados pelo usuário e salva no histórico (ProcessamentoRDO).
    """
    data = request.get_json()
    if not data or "resultados" not in data:
        return jsonify({"error": "Dados inválidos."}), 400

    resultados = data["resultados"]
    pdfs = data.get("pdfs", [])
    permissoes_input = data.get("permissoes", [])

    def _parse_dt_pte(s):
        """Parse 'DD/MM/YYYY H:MM:SS' or 'DD/MM/YYYY HH:MM:SS'."""
        s = s.strip()
        parts = s.split(' ')
        if len(parts) == 2:
            date_p, time_p = parts
            h, m, sec = time_p.split(':')
            s = f"{date_p} {h.zfill(2)}:{m}:{sec}"
        return datetime.strptime(s, "%d/%m/%Y %H:%M:%S")

    start_min = None
    end_max = None
    datas_encontradas = set()
    total = 0

    for key, colabs in resultados.items():
        total += len(colabs)
        partes = key.split('|')
        data_str = partes[0] if len(partes) > 0 else 'Sem Data'
        if data_str and data_str != 'Sem Data':
            datas_encontradas.add(data_str)

        ini_str = partes[1] if len(partes) > 1 else ''
        fim_str = partes[2] if len(partes) > 2 else ''

        if ini_str and ini_str.strip():
            try:
                dt_ini = _parse_dt_pte(ini_str)
                if not start_min or dt_ini < start_min:
                    start_min = dt_ini
            except (ValueError, Exception):
                pass
        if fim_str and fim_str.strip():
            try:
                dt_fim = _parse_dt_pte(fim_str)
                if not end_max or dt_fim > end_max:
                    end_max = dt_fim
            except (ValueError, Exception):
                pass

    ds = ", ".join(sorted(datas_encontradas)) if datas_encontradas else "Sem Data"

    import json as _json
    db = SessionLocal()
    try:
        ini_h = start_min.strftime('%H:%M') if start_min else None
        fim_h = end_max.strftime('%H:%M') if end_max else None
        proc = ProcessamentoRDO(
            nome_arquivo=ds,  # temporary, updated after commit with ID
            status="confirmado",
            total_nomes_extraidos=total,
            total_matches_auto=total,
            total_matches_revisao=0,
            total_sem_match=0,
            resultado_json=_json.dumps(resultados, ensure_ascii=False),
            pdfs_json=_json.dumps(pdfs, ensure_ascii=False) if pdfs else None,
            inicio_horario=ini_h,
            fim_horario=fim_h,
        )
        db.add(proc)
        db.commit()

        # Now that proc.id is set, update nome_arquivo with date+time from PDF
        if start_min and end_max:
            if start_min.date() == end_max.date():
                nome_final = f"{start_min.strftime('%d/%m/%Y')} {start_min.strftime('%H:%M')}–{end_max.strftime('%H:%M')}"
            else:
                nome_final = f"{start_min.strftime('%d/%m/%Y %H:%M')} a {end_max.strftime('%d/%m/%Y %H:%M')}"
        else:
            datas_sorted = sorted(datas_encontradas)
            if len(datas_sorted) == 1:
                nome_final = datas_sorted[0]
            elif datas_sorted:
                nome_final = f"{datas_sorted[0]} a {datas_sorted[-1]}"
            else:
                nome_final = "Sem Data"
        nome_final += f" — PTe #{proc.id}"
        proc.nome_arquivo = nome_final
        db.commit()

        # Salvar Permissões de Trabalho extraídas
        data_doc_iso = start_min.strftime('%Y-%m-%d') if start_min else None
        for pt in permissoes_input:
            numero = str(pt.get("numero_pt") or "").strip()
            desc = str(pt.get("descricao") or "").strip()
            if numero or desc:
                pt_rec = PermissaoTrabalho(
                    processamento_id=proc.id,
                    numero_pt=numero,
                    descricao=desc,
                    data_documento=data_doc_iso,
                )
                db.add(pt_rec)
        db.commit()

        registrar_log(db, "success", "pte", f"Confirmação de PTE: {total} colaboradores salvos no histórico.")

        return jsonify({"success": True})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/rdo/historico/<int:proc_id>/pdf/<string:filename>", methods=["GET"])
def download_pdf(proc_id, filename):
    """Serve um PDF armazenado vinculado a um processamento."""
    # Sanitize filename to prevent path traversal
    import re as _re2
    if not _re2.match(r'^[\w\-\.]+\.pdf$', filename, _re2.IGNORECASE):
        return jsonify({"error": "Nome de arquivo inválido"}), 400

    db = SessionLocal()
    try:
        proc = db.query(ProcessamentoRDO).filter(ProcessamentoRDO.id == proc_id).first()
        if not proc:
            return jsonify({"error": "Processamento não encontrado"}), 404

        try:
            pdfs = json.loads(proc.pdfs_json or '[]')
        except Exception:
            pdfs = []

        if filename not in pdfs:
            return jsonify({"error": "Arquivo não associado a este processamento"}), 403

        pdf_path = PDFS_DIR / filename
        if not pdf_path.exists():
            return jsonify({"error": "Arquivo não encontrado no servidor"}), 404

        return send_from_directory(str(PDFS_DIR), filename, as_attachment=True)
    finally:
        db.close()


# ─────────────────────────────────────────────
# RDO-OBRA (Relatorio Diario de Obra)
# ─────────────────────────────────────────────

def _obter_clima_data_especifica(data_iso: str) -> dict:
    """Obtém dados climáticos reais para uma data específica via Open-Meteo."""
    import requests as _rq
    from datetime import date as _dt_date

    conf = ler_clima_config() if False else {}
    # Tenta ler config; se falhar usa coordenadas padrão
    try:
        from pathlib import Path as _Path
        import json as _js
        cfg_path = _Path(__file__).resolve().parent.parent.parent / 'data' / 'clima_settings.json'
        if cfg_path.exists():
            with open(cfg_path, 'r', encoding='utf-8') as _f:
                conf = _js.load(_f)
    except Exception:
        pass

    lat = conf.get('lat', -20.3297)
    lon = conf.get('lon', -40.2925)

    try:
        today = _dt_date.today()
        target = _dt_date.fromisoformat(data_iso)
        days_ago = (today - target).days

        if days_ago > 5:
            url = "https://archive-api.open-meteo.com/v1/archive"
        else:
            url = "https://api.open-meteo.com/v1/forecast"

        resp = _rq.get(url, params={
            "latitude": lat, "longitude": lon,
            "start_date": data_iso, "end_date": data_iso,
            "hourly": "weathercode,precipitation",
            "timezone": "America/Sao_Paulo"
        }, timeout=10)
        resp.raise_for_status()
        hourly = resp.json().get("hourly", {})
        times = hourly.get("time", [])
        codes = hourly.get("weathercode", [])
        precip = hourly.get("precipitation", [])
    except Exception:
        return {"manha": "Bom", "tarde": "Bom", "noite": "Bom", "precipitacao": "0.0"}

    def _cat(code):
        if code is None: return "Bom"
        c = int(code)
        if c <= 2: return "Bom"
        elif c <= 48: return "Nublado"
        return "Chuvoso"

    def _worst(cats):
        if "Chuvoso" in cats: return "Chuvoso"
        if "Nublado" in cats: return "Nublado"
        return "Bom"

    manha, tarde, noite = [], [], []
    total_precip = 0.0
    for i, t in enumerate(times):
        try:
            hour = int(t.split('T')[1][:2]) if 'T' in t else -1
        except Exception:
            continue
        code = codes[i] if i < len(codes) else 0
        prec = float(precip[i] or 0) if i < len(precip) else 0.0
        total_precip += prec
        if 6 <= hour < 12:
            manha.append(_cat(code))
        elif 12 <= hour < 18:
            tarde.append(_cat(code))
        elif 18 <= hour <= 23:
            noite.append(_cat(code))

    return {
        "manha": _worst(manha) if manha else "Bom",
        "tarde": _worst(tarde) if tarde else "Bom",
        "noite": _worst(noite) if noite else "Bom",
        "precipitacao": str(round(total_precip, 1)),
    }


@api.route("/rdo-obra/dados", methods=["GET"])
def rdo_obra_dados():
    """Monta dados completos para o Relatorio Diario de Obra de uma data."""
    data_iso = request.args.get("data", "").strip()  # YYYY-MM-DD
    projeto_id = request.args.get("projeto_id", type=int)

    if not data_iso:
        return jsonify({"error": "Parâmetro 'data' obrigatório (YYYY-MM-DD)"}), 400

    from datetime import date as _dt_d
    try:
        data_obj = _dt_d.fromisoformat(data_iso)
    except ValueError:
        return jsonify({"error": "Data inválida"}), 400

    data_br = data_obj.strftime('%d/%m/%Y')

    db = SessionLocal()
    try:
        # 1. Buscar ProcessamentoRDO confirmado da data
        procs = db.query(ProcessamentoRDO).filter(
            ProcessamentoRDO.status == "confirmado"
        ).order_by(ProcessamentoRDO.data_processamento.desc()).all()

        proc_do_dia = None
        colabs_do_dia = []
        inicio_horario = None
        fim_horario = None

        for proc in procs:
            if not proc.resultado_json:
                continue
            try:
                resultado = json.loads(proc.resultado_json)
            except Exception:
                continue
            for key in resultado.keys():
                if key.startswith(data_br):
                    proc_do_dia = proc
                    inicio_horario = proc.inicio_horario
                    fim_horario = proc.fim_horario
                    seen_cpfs = set()
                    for colab in resultado[key]:
                        cpf_k = colab.get('cpf') or colab.get('nome', '')
                        if cpf_k not in seen_cpfs:
                            seen_cpfs.add(cpf_k)
                            colabs_do_dia.append(colab)
                    break
            if proc_do_dia:
                break

        # 2. Atividades do cronograma para a data
        atividades = []
        if projeto_id:
            tarefas = db.query(Tarefa).filter(
                Tarefa.projeto_id == projeto_id,
                Tarefa.inicio_previsto <= data_iso,
                Tarefa.fim_previsto >= data_iso,
                Tarefa.nivel >= 1,
            ).order_by(Tarefa.ordem).all()
            atividades = [t.to_dict() for t in tarefas]

        # 3. Permissões de Trabalho da data
        permissoes = []
        if proc_do_dia:
            pts = db.query(PermissaoTrabalho).filter(
                PermissaoTrabalho.processamento_id == proc_do_dia.id
            ).all()
            permissoes = [pt.to_dict() for pt in pts]
        # Também busca por data_documento
        if not permissoes:
            pts_by_date = db.query(PermissaoTrabalho).filter(
                PermissaoTrabalho.data_documento == data_iso
            ).all()
            permissoes = [pt.to_dict() for pt in pts_by_date]

        # 4. Clima para a data
        clima = _obter_clima_data_especifica(data_iso)

        return jsonify({
            "data": data_iso,
            "data_br": data_br,
            "processamento_id": proc_do_dia.id if proc_do_dia else None,
            "horarios": {
                "inicio_atividade": inicio_horario or "",
                "fim_atividade": fim_horario or "",
                "inicio_intervalo": "12:00",
                "fim_intervalo": "13:00",
            },
            "clima": clima,
            "efetivo": colabs_do_dia,
            "atividades": atividades,
            "permissoes": permissoes,
        })
    finally:
        db.close()


@api.route("/seguranca/permissoes", methods=["GET"])
def listar_permissoes():
    """Lista todas as Permissões de Trabalho registradas."""
    db = SessionLocal()
    try:
        data_filtro = request.args.get("data", "").strip()  # YYYY-MM-DD
        q = db.query(PermissaoTrabalho).order_by(PermissaoTrabalho.criado_em.desc())
        if data_filtro:
            q = q.filter(PermissaoTrabalho.data_documento == data_filtro)
        pts = q.limit(200).all()
        return jsonify({"permissoes": [pt.to_dict() for pt in pts]})
    finally:
        db.close()


# ─────────────────────────────────────────────
# CLIMA
# ─────────────────────────────────────────────

def ler_clima_config():
    config_path = Path(current_app.root_path).parent / 'data' / 'clima_settings.json'
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"cidade": "Vila Velha", "estado": "ES", "lat": -20.3297, "lon": -40.2925}

def salvar_clima_config(dados):
    config_path = Path(current_app.root_path).parent / 'data' / 'clima_settings.json'
    config_path.parent.mkdir(parents=True, exist_ok=True)
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False)

@api.route("/clima/config", methods=["GET"])
def get_clima_config():
    return jsonify(ler_clima_config())

@api.route("/clima/config", methods=["PUT"])
def update_clima_config():
    data = request.json
    cidade = data.get("cidade", "").strip()
    estado = data.get("estado", "").strip()
    
    if not cidade or not estado:
        return jsonify({"error": "Cidade e estado são obrigatórios"}), 400
        
    # Busca lat/lon
    from app.services.clima_service import buscar_coordenadas
    coords = buscar_coordenadas(cidade, estado)
    if not coords:
        return jsonify({"error": "Não foi possível encontrar as coordenadas para esta localização"}), 404
        
    config = {
        "cidade": coords['nome'],
        "estado": coords['estado'],
        "lat": coords['lat'],
        "lon": coords['lon']
    }
    salvar_clima_config(config)
    return jsonify({"success": True, "config": config})

@api.route("/clima", methods=["GET"])
def get_clima():
    """Retorna dados climáticos com configuração dinâmica."""
    forcar = request.args.get("forcar", "false").lower() == "true"
    conf = ler_clima_config()
    dados = obter_clima(cidade=conf['cidade'], estado=conf['estado'], lat=conf['lat'], lon=conf['lon'], forcar_atualizacao=forcar)
    return jsonify(dados)


# ─────────────────────────────────────────────
# SISTEMA / DASHBOARD
# ─────────────────────────────────────────────

@api.route("/dashboard/stats", methods=["GET"])
def dashboard_stats():
    """Estatísticas gerais para o dashboard."""
    db = SessionLocal()
    try:
        total_colabs = db.query(Colaborador).filter(Colaborador.ativo == True).count()
        total_mod = db.query(Colaborador).filter(Colaborador.ativo == True, Colaborador.categoria == 'MOD').count()
        total_moi = db.query(Colaborador).filter(Colaborador.ativo == True, Colaborador.categoria == 'MOI').count()
        total_processamentos = db.query(ProcessamentoRDO).count()
        ultimo_proc = db.query(ProcessamentoRDO).filter(
            ProcessamentoRDO.status == "confirmado"
        ).order_by(ProcessamentoRDO.data_processamento.desc()).first()
        total_vinculos = db.query(Vinculo).filter(Vinculo.confirmado == True).count()
        total_projetos = db.query(Projeto).filter(Projeto.status == 'ativo').count()
        total_pts = db.query(PermissaoTrabalho).count()

        logs_recentes = db.query(LogAtividade).order_by(LogAtividade.data.desc()).limit(10).all()

        # Efetivo do último PTe processado
        efetivo_ultimo = []
        if ultimo_proc and ultimo_proc.resultado_json:
            try:
                res = json.loads(ultimo_proc.resultado_json)
                for key, colabs in res.items():
                    for c in colabs:
                        if not any(x.get('cpf') == c.get('cpf') for x in efetivo_ultimo):
                            efetivo_ultimo.append(c)
            except Exception:
                pass

        return jsonify({
            "total_colaboradores": total_colabs,
            "total_mod": total_mod,
            "total_moi": total_moi,
            "total_processamentos": total_processamentos,
            "total_vinculos_confirmados": total_vinculos,
            "total_projetos_ativos": total_projetos,
            "total_permissoes": total_pts,
            "ultimo_processamento": ultimo_proc.to_dict() if ultimo_proc else None,
            "efetivo_ultimo_pte": efetivo_ultimo[:5],
            "logs_recentes": [l.to_dict() for l in logs_recentes]
        })
    finally:
        db.close()


# ─────────────────────────────────────────────
# PLANEJAMENTO DE OBRAS (Project Mirror)
# ─────────────────────────────────────────────

import re as _re2
import csv
import io as _io
from datetime import date as _date, timedelta as _td

_MESES_BR = {
    'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,
    'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12,
    'janeiro':1,'fevereiro':2,'março':3,'marco':3,'abril':4,'maio':5,
    'junho':6,'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12,
}

_TIME_SUFFIX = _re2.compile(r'\s+\d{1,2}:\d{2}(?::\d{2})?$')

def _parse_data_br(val):
    if not val: return None
    s = _TIME_SUFFIX.sub('', str(val).strip())   # remove " 09:00" ou " 09:00:00"
    if not s or s in ('-','NA','N/A','#','0'): return None
    if _re2.match(r'\d{4}-\d{2}-\d{2}', s): return s[:10]
    m = _re2.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', s)
    if m: return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    m = _re2.match(r'^(\d{1,2})[\s/\-]+(\w+)[\s/\-]+(\d{4})$', s, _re2.I)
    if m:
        mes = _MESES_BR.get(m.group(2).lower().replace('ç','c').replace('é','e').replace('ê','e').replace('ã','a'))
        if mes: return f"{m.group(3)}-{mes:02d}-{int(m.group(1)):02d}"
    return None

def _parse_pct(val):
    if not val: return 0.0
    s = str(val).strip().rstrip('%').replace(',','.')
    try: return min(100.0, max(0.0, float(s)))
    except: return 0.0

def _parse_int(val):
    if not val: return 0
    m = _re2.match(r'(\d+)', str(val).strip())
    return int(m.group(1)) if m else 0

def _calcular_critico(tarefas):
    """Marca tarefas no caminho crítico (tarefa mais tarde do projeto)."""
    fins = [t.fim_previsto for t in tarefas if t.fim_previsto and not t.is_marco]
    if not fins: return set()
    fim_projeto = max(fins)
    criticos = set()
    # Marca as últimas e rastreia predecessoras
    fila = [t for t in tarefas if t.fim_previsto == fim_projeto]
    visitados = set()
    task_by_cod = {t.codigo: t for t in tarefas if t.codigo}
    task_by_id = {t.id: t for t in tarefas}
    while fila:
        t = fila.pop()
        if t.id in visitados: continue
        visitados.add(t.id)
        criticos.add(t.id)
        if t.predecessoras:
            try:
                preds = json.loads(t.predecessoras)
                for p in preds:
                    pred_t = task_by_cod.get(str(p))
                    if not pred_t:
                        try: pred_t = task_by_id.get(int(p))
                        except: pass
                    if pred_t: fila.append(pred_t)
            except: pass
    return criticos


@api.route("/projetos", methods=["GET"])
def listar_projetos():
    db = SessionLocal()
    try:
        projetos = db.query(Projeto).order_by(Projeto.criado_em.desc()).all()
        return jsonify({"projetos": [p.to_dict() for p in projetos]})
    finally:
        db.close()


@api.route("/projetos", methods=["POST"])
def criar_projeto():
    db = SessionLocal()
    try:
        d = request.get_json(force=True) or {}
        nome = str(d.get("nome","")).strip()
        if not nome: return jsonify({"error":"Nome obrigatório"}), 400
        p = Projeto(nome=nome, descricao=d.get("descricao",""))
        db.add(p); db.commit()
        return jsonify({"success": True, "id": p.id, "projeto": p.to_dict()})
    except Exception as e:
        db.rollback(); return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/projetos/<int:pid>", methods=["GET"])
def detalhe_projeto(pid):
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p: return jsonify({"error":"Projeto não encontrado"}), 404
        criticos = _calcular_critico(p.tarefas)
        tarefas = []
        for t in p.tarefas:
            d = t.to_dict()
            d["is_critico"] = t.id in criticos
            tarefas.append(d)
        return jsonify({"projeto": p.to_dict(), "tarefas": tarefas})
    finally:
        db.close()


@api.route("/projetos/<int:pid>", methods=["DELETE"])
def deletar_projeto(pid):
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p: return jsonify({"error":"Não encontrado"}), 404
        db.delete(p); db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


@api.route("/projetos/<int:pid>/importar", methods=["POST"])
def importar_cronograma(pid):
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p: return jsonify({"error":"Projeto não encontrado"}), 404

        if "file" not in request.files:
            return jsonify({"error":"Nenhum arquivo enviado"}), 400
        f = request.files["file"]
        ext = Path(f.filename).suffix.lower()

        rows = []
        if ext in ('.xlsx', '.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(c.value or '').strip() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
        elif ext in ('.csv', '.txt', '.tsv'):
            content = f.read().decode('utf-8-sig', errors='replace')
            # Auto-detecta delimitador: TSV (MS Project) ou CSV
            sample = content[:3000]
            tabs = sample.count('\t'); commas = sample.count(','); semis = sample.count(';')
            delim = '\t' if tabs >= max(commas, semis) else (',' if commas >= semis else ';')
            reader = csv.DictReader(_io.StringIO(content), delimiter=delim)
            rows = [dict(r) for r in reader]
        else:
            return jsonify({"error":"Formato inválido. Use .csv, .xlsx, .xls ou .txt"}), 400

        if not rows: return jsonify({"error":"Arquivo vazio"}), 400

        # Detecta colunas (case-insensitive, ignora acentos)
        def _norm(s): return s.lower().strip().replace('ç','c').replace('ã','a').replace('á','a').replace('é','e').replace('ê','e').replace('í','i').replace('ó','o').replace('ô','o').replace('ú','u')
        def col(row, *keys):
            # Prioridade: itera chaves de busca em ordem para evitar falsos positivos
            for key in keys:
                for k in row:
                    kl = _norm(k)
                    if key == kl or key in kl: return row[k]
            return ''

        # Limpa tarefas existentes do projeto
        db.query(Tarefa).filter(Tarefa.projeto_id == pid).delete()
        db.commit()

        importadas = 0
        for i, row in enumerate(rows):
            # Ignora linhas inativas (MS Project)
            ativo = col(row,'ativo','active')
            if ativo and str(ativo).strip().lower() in ('não','nao','no','false','0'):
                continue

            nome = col(row,'nome','atividade','descricao','task name','task','name')
            if not nome or not nome.strip() or nome.strip() in ('-',''):
                continue

            ini_prev = _parse_data_br(col(row,'inicio','start','data ini','dt ini'))
            fim_prev = _parse_data_br(col(row,'termino','termino','fim','end','data fim','dt fim','finish'))
            ini_real = _parse_data_br(col(row,'inicio real','real start','data real ini'))
            fim_real = _parse_data_br(col(row,'termino real','real end','data real fim'))
            progresso = _parse_pct(col(row,'% conc','%conc','progresso','progress','concluido','% concl','avanco'))
            dur_raw = col(row,'duracao','dur','duration')
            duracao = _parse_int(dur_raw) if dur_raw else None
            predecessoras = col(row,'predecess','predeces','predecessor')
            # Limpa sufixos MS Project: "2II", "20TI+3d", "43II+11d" → extrai só o número
            pred_list = []
            if predecessoras and predecessoras.strip():
                for item in _re2.split(r'[,;]', predecessoras):
                    m_pred = _re2.match(r'(\d+)', item.strip())
                    if m_pred: pred_list.append(m_pred.group(1))
            mo = _parse_int(col(row,'mao de obra','mo ','m.o','recurso mo','h.h','hh'))
            eq = _parse_int(col(row,'equipamento','equip','maquina'))
            # Código: coluna Id do MS Project
            codigo = col(row,'id','codigo','wbs','n.','item','seq')
            if not codigo or not str(codigo).strip() or str(codigo).strip() in ('None',''):
                codigo = str(i)
            # Nível hierárquico: "Nível da estrutura de tópicos" ou "Level"
            nivel_raw = col(row,'nivel da estrutura','nivel','level','hierarquia','topicos','outline')
            nivel = _parse_int(nivel_raw) if nivel_raw else 0

            # Calcula duração a partir das datas se não informada
            if not duracao and ini_prev and fim_prev:
                try:
                    d1 = _date.fromisoformat(ini_prev)
                    d2 = _date.fromisoformat(fim_prev)
                    duracao = max(1, (d2 - d1).days + 1)
                except: pass

            t = Tarefa(
                projeto_id=pid, codigo=str(codigo).strip(), nome=nome.strip(),
                nivel=nivel, ordem=i,
                duracao=duracao,
                inicio_previsto=ini_prev, fim_previsto=fim_prev,
                inicio_real=ini_real, fim_real=fim_real,
                progresso=progresso,
                predecessoras=json.dumps(pred_list) if pred_list else None,
                recursos_mo=mo, recursos_eq=eq,
                peso=float(duracao or 1),
            )
            db.add(t)
            importadas += 1

        db.commit()
        registrar_log(db, "success", "planejamento", f"Cronograma importado: {importadas} tarefas", f"Projeto: {p.nome}")
        return jsonify({"success": True, "importadas": importadas})
    except Exception as e:
        db.rollback(); return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/tarefas/<int:tid>", methods=["PUT"])
def atualizar_tarefa(tid):
    db = SessionLocal()
    try:
        t = db.query(Tarefa).get(tid)
        if not t: return jsonify({"error":"Não encontrada"}), 404
        d = request.get_json(force=True) or {}
        for campo in ('progresso','inicio_real','fim_real','inicio_previsto','fim_previsto','recursos_mo','recursos_eq'):
            if campo in d:
                val = d[campo]
                if campo == 'progresso': val = min(100.0, max(0.0, float(val or 0)))
                setattr(t, campo, val)
        db.commit()
        return jsonify({"success": True, "tarefa": t.to_dict()})
    except Exception as e:
        db.rollback(); return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/projetos/<int:pid>/curva-s", methods=["GET"])
def curva_s(pid):
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p: return jsonify({"error":"Não encontrado"}), 404
        tarefas = [t for t in p.tarefas if t.inicio_previsto and t.fim_previsto]
        if not tarefas: return jsonify({"labels":[], "previsto":[], "real":[]})

        peso_total = sum(t.peso or 1 for t in tarefas)
        all_dates = [t.inicio_previsto for t in tarefas] + [t.fim_previsto for t in tarefas]
        d_min = _date.fromisoformat(min(all_dates))
        d_max = _date.fromisoformat(max(all_dates))
        hoje = _date.today()

        labels, previsto_list, real_list = [], [], []
        d = d_min
        while d <= d_max + _td(days=7):
            prev_acc = 0.0
            real_acc = 0.0
            for t in tarefas:
                try:
                    di = _date.fromisoformat(t.inicio_previsto)
                    df = _date.fromisoformat(t.fim_previsto)
                    dur = max(1, (df - di).days)
                    p_prog = min(1.0, max(0.0, (d - di).days / dur)) if d >= di else 0.0
                    prev_acc += (t.peso or 1) * p_prog
                    if d <= hoje:
                        if t.inicio_real:
                            ri = _date.fromisoformat(t.inicio_real)
                            if d >= ri:
                                real_acc += (t.peso or 1) * (t.progresso or 0) / 100.0
                except: pass
            labels.append(d.isoformat())
            previsto_list.append(round(prev_acc / peso_total * 100, 2))
            real_list.append(round(real_acc / peso_total * 100, 2) if d <= hoje else None)
            d += _td(days=7)

        return jsonify({"labels": labels, "previsto": previsto_list, "real": real_list})
    finally:
        db.close()


@api.route("/projetos/<int:pid>/curva-s-semanal", methods=["GET"])
def curva_s_semanal(pid):
    """Retorna breakdown semanal da Curva S: Semana, Período, Previsto Sem%, Previsto Ac%, Real Sem%, Real Ac%, Desvio Ac%."""
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p:
            return jsonify({"error": "Não encontrado"}), 404
        tarefas = [t for t in p.tarefas if t.inicio_previsto and t.fim_previsto]
        if not tarefas:
            return jsonify({"semanas": []})

        peso_total = sum(t.peso or 1 for t in tarefas)
        all_dates = [t.inicio_previsto for t in tarefas] + [t.fim_previsto for t in tarefas]
        d_min = _date.fromisoformat(min(all_dates))
        d_max = _date.fromisoformat(max(all_dates))
        hoje = _date.today()

        # Align to Monday
        wd = d_min.weekday()
        seg_ini = d_min - _td(days=wd)

        def _prev_ac_at(d):
            acc = 0.0
            for t in tarefas:
                try:
                    di = _date.fromisoformat(t.inicio_previsto)
                    df = _date.fromisoformat(t.fim_previsto)
                    dur = max(1, (df - di).days)
                    prog = min(1.0, max(0.0, (d - di).days / dur)) if d >= di else 0.0
                    acc += (t.peso or 1) * prog
                except Exception:
                    pass
            return round(acc / peso_total * 100, 2)

        def _real_ac_at(d):
            if d > hoje:
                return None
            acc = 0.0
            for t in tarefas:
                try:
                    if t.inicio_real:
                        ri = _date.fromisoformat(t.inicio_real)
                        if d >= ri:
                            acc += (t.peso or 1) * (t.progresso or 0) / 100.0
                except Exception:
                    pass
            return round(acc / peso_total * 100, 2)

        semanas = []
        seg = seg_ini
        num_sem = 1
        prev_prev_ac = 0.0
        prev_real_ac = 0.0
        while seg <= d_max:
            dom = seg + _td(days=6)
            prev_ac = _prev_ac_at(dom)
            real_ac = _real_ac_at(dom)
            prev_sem = round(prev_ac - prev_prev_ac, 2)
            real_sem = round((real_ac - prev_real_ac), 2) if real_ac is not None else None
            desvio_ac = round((real_ac or 0) - prev_ac, 2) if real_ac is not None else None
            semanas.append({
                "semana": num_sem,
                "seg": seg.isoformat(),
                "dom": dom.isoformat(),
                "previsto_sem": prev_sem,
                "previsto_ac": prev_ac,
                "real_sem": real_sem,
                "real_ac": real_ac,
                "desvio_ac": desvio_ac,
            })
            prev_prev_ac = prev_ac
            if real_ac is not None:
                prev_real_ac = real_ac
            seg += _td(days=7)
            num_sem += 1

        return jsonify({"semanas": semanas, "peso_total": peso_total})
    finally:
        db.close()


@api.route("/projetos/<int:pid>/histograma", methods=["GET"])
def histograma(pid):
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p: return jsonify({"error":"Não encontrado"}), 404
        tarefas = [t for t in p.tarefas if t.inicio_previsto and t.fim_previsto and t.nivel >= 1]
        if not tarefas: return jsonify({"labels":[], "mo_prev":[], "mo_real":[], "eq_prev":[], "eq_real":[]})

        all_dates = [t.inicio_previsto for t in tarefas] + [t.fim_previsto for t in tarefas]
        d_min = _date.fromisoformat(min(all_dates))
        d_max = _date.fromisoformat(max(all_dates))

        labels, mo_prev, mo_real, eq_prev, eq_real = [], [], [], [], []
        d = d_min
        while d <= d_max:
            d_end = d + _td(days=6)
            lbl = f"{d.strftime('%d/%m')}"
            mp, mr, ep, er = 0, 0, 0, 0
            for t in tarefas:
                try:
                    di = _date.fromisoformat(t.inicio_previsto)
                    df = _date.fromisoformat(t.fim_previsto)
                    if di <= d_end and df >= d:
                        mp += t.recursos_mo or 0
                        ep += t.recursos_eq or 0
                    if t.inicio_real:
                        ri = _date.fromisoformat(t.inicio_real)
                        rf = _date.fromisoformat(t.fim_real) if t.fim_real else _date.today()
                        if ri <= d_end and rf >= d:
                            mr += t.recursos_mo or 0
                            er += t.recursos_eq or 0
                except: pass
            labels.append(lbl); mo_prev.append(mp); mo_real.append(mr)
            eq_prev.append(ep); eq_real.append(er)
            d += _td(days=7)

        return jsonify({"labels": labels, "mo_prev": mo_prev, "mo_real": mo_real,
                        "eq_prev": eq_prev, "eq_real": eq_real})
    finally:
        db.close()


@api.route("/projetos/<int:pid>/exportar", methods=["GET"])
def exportar_cronograma(pid):
    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p: return jsonify({"error":"Não encontrado"}), 404
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cronograma"
        verde = "1A6B3C"
        thin = Side(style="thin", color="CCCCCC")
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)
        headers = ["Cód.","Atividade","Nível","Duração","Início Prev.","Fim Prev.","Início Real","Fim Real","% Concl.","MO","Equip.","Predecessoras"]
        widths  = [10, 50, 8, 10, 14, 14, 14, 14, 10, 8, 8, 20]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=verde)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borda
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        for ri, t in enumerate(p.tarefas, 2):
            preds = ', '.join(json.loads(t.predecessoras)) if t.predecessoras else ''
            fill_c = "EAF4EE" if ri % 2 == 0 else "FFFFFF"
            row_data = [t.codigo, t.nome, t.nivel, t.duracao,
                        t.inicio_previsto, t.fim_previsto, t.inicio_real, t.fim_real,
                        f"{t.progresso:.1f}%", t.recursos_mo, t.recursos_eq, preds]
            indent = (t.nivel or 0) * 4
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=fill_c)
                cell.border = borda
                cell.alignment = Alignment(vertical="center", indent=(indent if ci == 2 else 0))
            if t.nivel == 0:
                for ci in range(1, len(headers)+1):
                    ws.cell(row=ri, column=ci).font = Font(bold=True, size=10)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f"cronograma_{p.nome.replace(' ','_')}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    finally:
        db.close()


@api.route("/projetos/<int:pid>/salvar-editor", methods=["POST"])
def salvar_editor(pid):
    """Salva em lote todas as tarefas do editor inline."""
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p:
            return jsonify({"error": "Projeto não encontrado"}), 404
        dados = request.get_json(force=True) or {}
        tarefas_data = dados.get("tarefas", [])

        # Substitui todas as tarefas do projeto
        db.query(Tarefa).filter(Tarefa.projeto_id == pid).delete()
        for i, td in enumerate(tarefas_data):
            preds = td.get("predecessoras") or []
            if isinstance(preds, str):
                preds = [x.strip() for x in _re2.split(r'[,;]', preds) if x.strip()]
            dur = td.get("duracao")
            try: dur = int(dur) if dur else None
            except: dur = None
            t = Tarefa(
                projeto_id=pid,
                codigo=str(td.get("codigo") or "").strip() or None,
                nome=str(td.get("nome") or "").strip() or "Sem nome",
                nivel=int(td.get("nivel") or 0),
                ordem=int(td.get("ordem", i)),
                duracao=dur,
                inicio_previsto=td.get("inicio_previsto") or None,
                fim_previsto=td.get("fim_previsto") or None,
                inicio_real=td.get("inicio_real") or None,
                fim_real=td.get("fim_real") or None,
                progresso=min(100.0, max(0.0, float(td.get("progresso") or 0))),
                predecessoras=json.dumps(preds) if preds else None,
                recursos_mo=int(td.get("recursos_mo") or 0),
                recursos_eq=int(td.get("recursos_eq") or 0),
                peso=float(dur or 1),   # peso = duração para Curva S ponderada por tempo
                is_marco=bool(td.get("is_marco")),
            )
            db.add(t)
        db.commit()
        registrar_log(db, "success", "planejamento", f"Editor: {len(tarefas_data)} tarefas salvas", f"Projeto: {p.nome}")
        return jsonify({"success": True, "salvas": len(tarefas_data)})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


@api.route("/projetos/<int:pid>/modelo-csv", methods=["GET"])
def modelo_csv(pid):
    from flask import Response
    linhas = [
        "ID\tNome\tDuração\tInício\tTérmino\t% Concluído\tPredecessoras\tNível da estrutura de tópicos\tMO\tEquipamentos",
        "1\tMobilização\t5 dias\t06/01/2026\t10/01/2026\t100%\t\t1\t5\t1",
        "2\tTerraplenagem\t20 dias\t12/01/2026\t06/02/2026\t50%\t1\t1\t15\t3",
        "3\t   Preparo do terreno\t10 dias\t12/01/2026\t23/01/2026\t60%\t1\t2\t8\t2",
        "4\t   Aterro\t10 dias\t26/01/2026\t06/02/2026\t40%\t3\t2\t7\t1",
        "5\tFundação\t30 dias\t09/02/2026\t20/03/2026\t0%\t2\t1\t20\t2",
    ]
    return Response('\n'.join(linhas), mimetype='text/tab-separated-values',
                    headers={"Content-Disposition": "attachment; filename=modelo_cronograma.tsv"})


@api.route("/projetos/<int:pid>/importar-xml", methods=["POST"])
def importar_xml(pid):
    """Importa cronograma a partir de XML do MS Project (.xml)."""
    import xml.etree.ElementTree as _ET
    db = SessionLocal()
    try:
        p = db.query(Projeto).get(pid)
        if not p:
            return jsonify({"error": "Projeto não encontrado"}), 404
        if "file" not in request.files:
            return jsonify({"error": "Nenhum arquivo enviado"}), 400

        content = request.files["file"].read()
        try:
            root = _ET.fromstring(content)
        except _ET.ParseError as e:
            return jsonify({"error": f"XML inválido: {e}"}), 400

        # Ignora namespace — funciona com qualquer versão do MS Project
        def _sn(tag):
            return tag.split('}')[-1] if '}' in tag else tag

        def _find(el, name):
            for ch in el:
                if _sn(ch.tag) == name:
                    return ch
            return None

        def _findall(el, name):
            return [ch for ch in el if _sn(ch.tag) == name]

        def _txt(el, name, default=''):
            ch = _find(el, name)
            return (ch.text or default).strip() if ch is not None else default

        def _dur_days(s):
            """Converte duração ISO 8601 do MS Project (PT8H0M0S) em dias úteis (8h/dia)."""
            if not s: return None
            m = _re2.match(r'P(?:(\d+)D)?(?:T(?:(\d+(?:\.\d+)?)H)?(?:(\d+)M)?(?:[\d.]+S)?)?', s)
            if m:
                days = float(m.group(1) or 0)
                hours = float(m.group(2) or 0)
                total_h = days * 8 + hours
                d = max(1, round(total_h / 8)) if total_h > 0 else None
                return d
            return None

        def _dt(s):
            if not s or s.upper() in ('NA', 'N/A', '0', ''): return None
            return s[:10] if len(s) >= 10 else None

        tasks_el = _find(root, 'Tasks')
        if tasks_el is None:
            return jsonify({"error": "Tag <Tasks> não encontrada — verifique se é um XML do MS Project"}), 400

        # Construir mapa UID_tarefa → nome_recurso (empresa/responsável)
        resources_el = _find(root, 'Resources')
        uid_to_resource = {}
        if resources_el:
            for res_el in _findall(resources_el, 'Resource'):
                r_uid = _txt(res_el, 'UID')
                r_name = _txt(res_el, 'Name')
                if r_uid and r_name and r_uid != '0':
                    uid_to_resource[r_uid] = r_name

        assignments_el = _find(root, 'Assignments')
        task_uid_to_resource = {}
        if assignments_el:
            for asgn_el in _findall(assignments_el, 'Assignment'):
                t_uid = _txt(asgn_el, 'TaskUID')
                r_uid = _txt(asgn_el, 'ResourceUID')
                if t_uid and r_uid and r_uid in uid_to_resource:
                    # Keep first/primary resource per task
                    if t_uid not in task_uid_to_resource:
                        task_uid_to_resource[t_uid] = uid_to_resource[r_uid]

        # Primeiro passo: mapear UID → codigo WBS para resolver predecessoras
        uid_to_cod = {}
        tasks_raw = []
        for task_el in _findall(tasks_el, 'Task'):
            uid  = _txt(task_el, 'UID')
            tid  = _txt(task_el, 'ID')
            nome = _txt(task_el, 'Name')
            if not tid or tid == '0' or not nome:
                continue
            inactive = _txt(task_el, 'Inactive', '0')
            if inactive in ('1', 'true'):
                continue
            summary = _txt(task_el, 'Summary', '0')  # tarefas resumo
            # Usar WBS ou OutlineNumber como código EDT (ex: "1.2.3"), fallback p/ ID sequencial
            wbs = _txt(task_el, 'WBS') or _txt(task_el, 'OutlineNumber') or tid
            uid_to_cod[uid] = wbs
            tasks_raw.append((wbs, uid, nome, task_el, summary))

        # Limpa tarefas existentes
        db.query(Tarefa).filter(Tarefa.projeto_id == pid).delete()
        db.commit()

        importadas = 0
        for ordem, (wbs, uid, nome, task_el, summary) in enumerate(tasks_raw):
            outline = int(_txt(task_el, 'OutlineLevel') or '1')
            dur_str = _txt(task_el, 'Duration')
            dur = _dur_days(dur_str)
            ini_prev = _dt(_txt(task_el, 'Start'))
            fim_prev = _dt(_txt(task_el, 'Finish'))
            ini_real = _dt(_txt(task_el, 'ActualStart'))
            fim_real = _dt(_txt(task_el, 'ActualFinish'))
            pct       = float(_txt(task_el, 'PercentComplete') or '0')

            # Calcular duração pelas datas se não informada
            if not dur and ini_prev and fim_prev:
                try:
                    d1 = _date.fromisoformat(ini_prev)
                    d2 = _date.fromisoformat(fim_prev)
                    dur = max(1, (d2 - d1).days + 1)
                except: pass

            # Predecessoras: resolver UIDs → IDs
            pred_list = []
            for pl in _findall(task_el, 'PredecessorLink'):
                puid = _txt(pl, 'PredecessorUID')
                if puid and puid != '0' and puid in uid_to_cod:
                    pred_list.append(uid_to_cod[puid])

            t = Tarefa(
                projeto_id=pid, codigo=wbs, nome=nome,
                nivel=max(0, outline - 1),  # OutlineLevel 1 = nivel 0 (grupo)
                ordem=ordem,
                duracao=dur,
                inicio_previsto=ini_prev, fim_previsto=fim_prev,
                inicio_real=ini_real, fim_real=fim_real,
                progresso=min(100.0, max(0.0, pct)),
                predecessoras=json.dumps(pred_list) if pred_list else None,
                recursos_mo=0, recursos_eq=0,
                peso=float(dur or 1),
                is_marco=(summary == '0' and dur == 0),
                responsavel=task_uid_to_resource.get(uid),
            )
            db.add(t)
            importadas += 1

        db.commit()
        registrar_log(db, "success", "planejamento", f"XML importado: {importadas} tarefas", f"Projeto: {p.nome}")
        return jsonify({"success": True, "importadas": importadas})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()


# ─────────────────────────────────────────────
# CADASTROS BASE (Equipamentos, Veículos, Terceiros)
# ─────────────────────────────────────────────

@api.route("/equipamentos", methods=["GET"])
def listar_equipamentos():
    db = SessionLocal()
    try:
        q = request.args.get("q", "").strip()
        query = db.query(Equipamento)
        if q:
            query = query.filter(Equipamento.nome.ilike(f"%{q}%") | Equipamento.codigo.ilike(f"%{q}%"))
        return jsonify({"equipamentos": [e.to_dict() for e in query.order_by(Equipamento.nome).all()]})
    finally:
        db.close()

@api.route("/equipamentos", methods=["POST"])
def criar_equipamento():
    db = SessionLocal()
    try:
        d = request.get_json(force=True) or {}
        nome = str(d.get("nome","")).strip()
        if not nome: return jsonify({"error":"Nome obrigatório"}), 400
        e = Equipamento(nome=nome, codigo=d.get("codigo","").strip() or None,
                        status=d.get("status","ativo"))
        db.add(e); db.commit()
        return jsonify({"success": True, "equipamento": e.to_dict()})
    except Exception as ex:
        db.rollback(); return jsonify({"error": str(ex)}), 500
    finally:
        db.close()

@api.route("/equipamentos/<int:eid>", methods=["PUT"])
def atualizar_equipamento(eid):
    db = SessionLocal()
    try:
        e = db.query(Equipamento).get(eid)
        if not e: return jsonify({"error":"Não encontrado"}), 404
        d = request.get_json(force=True) or {}
        if "nome" in d: e.nome = str(d["nome"]).strip()
        if "codigo" in d: e.codigo = str(d["codigo"]).strip() or None
        if "status" in d: e.status = str(d["status"])
        db.commit()
        return jsonify({"success": True, "equipamento": e.to_dict()})
    except Exception as ex:
        db.rollback(); return jsonify({"error": str(ex)}), 500
    finally:
        db.close()

@api.route("/equipamentos/<int:eid>", methods=["DELETE"])
def deletar_equipamento(eid):
    db = SessionLocal()
    try:
        e = db.query(Equipamento).get(eid)
        if not e: return jsonify({"error":"Não encontrado"}), 404
        db.delete(e); db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


@api.route("/veiculos", methods=["GET"])
def listar_veiculos():
    db = SessionLocal()
    try:
        q = request.args.get("q", "").strip()
        query = db.query(Veiculo)
        if q:
            query = query.filter(Veiculo.placa.ilike(f"%{q}%") | Veiculo.modelo.ilike(f"%{q}%") | Veiculo.empresa.ilike(f"%{q}%"))
        return jsonify({"veiculos": [v.to_dict() for v in query.order_by(Veiculo.placa).all()]})
    finally:
        db.close()

@api.route("/veiculos", methods=["POST"])
def criar_veiculo():
    db = SessionLocal()
    try:
        d = request.get_json(force=True) or {}
        placa = str(d.get("placa","")).strip().upper()
        if not placa: return jsonify({"error":"Placa obrigatória"}), 400
        v = Veiculo(placa=placa, modelo=d.get("modelo","").strip() or None,
                    empresa=d.get("empresa","").strip() or None)
        db.add(v); db.commit()
        return jsonify({"success": True, "veiculo": v.to_dict()})
    except Exception as ex:
        db.rollback(); return jsonify({"error": str(ex)}), 500
    finally:
        db.close()

@api.route("/veiculos/<int:vid>", methods=["PUT"])
def atualizar_veiculo(vid):
    db = SessionLocal()
    try:
        v = db.query(Veiculo).get(vid)
        if not v: return jsonify({"error":"Não encontrado"}), 404
        d = request.get_json(force=True) or {}
        if "placa" in d: v.placa = str(d["placa"]).strip().upper()
        if "modelo" in d: v.modelo = str(d["modelo"]).strip() or None
        if "empresa" in d: v.empresa = str(d["empresa"]).strip() or None
        db.commit()
        return jsonify({"success": True, "veiculo": v.to_dict()})
    except Exception as ex:
        db.rollback(); return jsonify({"error": str(ex)}), 500
    finally:
        db.close()

@api.route("/veiculos/<int:vid>", methods=["DELETE"])
def deletar_veiculo(vid):
    db = SessionLocal()
    try:
        v = db.query(Veiculo).get(vid)
        if not v: return jsonify({"error":"Não encontrado"}), 404
        db.delete(v); db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


@api.route("/terceiros", methods=["GET"])
def listar_terceiros():
    db = SessionLocal()
    try:
        q = request.args.get("q", "").strip()
        query = db.query(HistoricoTerceiro)
        if q:
            query = query.filter(HistoricoTerceiro.nome.ilike(f"%{q}%") |
                                 HistoricoTerceiro.cpf.ilike(f"%{q}%") |
                                 HistoricoTerceiro.empresa.ilike(f"%{q}%"))
        items = query.order_by(HistoricoTerceiro.nome).all()
        return jsonify({"terceiros": [t.to_dict() for t in items]})
    finally:
        db.close()

@api.route("/terceiros", methods=["POST"])
def criar_terceiro():
    db = SessionLocal()
    try:
        d = request.get_json(force=True) or {}
        nome = _normalizar_nome(d.get("nome",""))
        if not nome: return jsonify({"error":"Nome obrigatório"}), 400
        cpf = _limpar_cpf(d.get("cpf","")) or None
        t = HistoricoTerceiro(nome=nome, cpf=cpf,
                              placa=d.get("placa","").strip().upper() or None,
                              empresa=d.get("empresa","").strip() or None,
                              local=d.get("local","").strip() or None,
                              motivo=d.get("motivo","").strip() or None)
        db.add(t); db.commit()
        return jsonify({"success": True, "terceiro": t.to_dict()})
    except Exception as ex:
        db.rollback(); return jsonify({"error": str(ex)}), 500
    finally:
        db.close()

@api.route("/terceiros/<int:tid>", methods=["PUT"])
def atualizar_terceiro(tid):
    db = SessionLocal()
    try:
        t = db.query(HistoricoTerceiro).get(tid)
        if not t: return jsonify({"error":"Não encontrado"}), 404
        d = request.get_json(force=True) or {}
        if "nome" in d: t.nome = _normalizar_nome(d["nome"])
        if "cpf" in d: t.cpf = _limpar_cpf(d["cpf"]) or None
        if "placa" in d: t.placa = str(d["placa"]).strip().upper() or None
        if "empresa" in d: t.empresa = str(d["empresa"]).strip() or None
        db.commit()
        return jsonify({"success": True, "terceiro": t.to_dict()})
    except Exception as ex:
        db.rollback(); return jsonify({"error": str(ex)}), 500
    finally:
        db.close()

@api.route("/terceiros/<int:tid>", methods=["DELETE"])
def deletar_terceiro(tid):
    db = SessionLocal()
    try:
        t = db.query(HistoricoTerceiro).get(tid)
        if not t: return jsonify({"error":"Não encontrado"}), 404
        db.delete(t); db.commit()
        return jsonify({"success": True})
    finally:
        db.close()


@api.route("/logs", methods=["GET"])
def get_logs():
    """Retorna logs do sistema."""
    db = SessionLocal()
    try:
        modulo = request.args.get("modulo")
        limite = int(request.args.get("limite", 50))

        query = db.query(LogAtividade).order_by(LogAtividade.data.desc())
        if modulo:
            query = query.filter(LogAtividade.modulo == modulo)
        logs = query.limit(limite).all()

        return jsonify({"logs": [l.to_dict() for l in logs]})
    finally:
        db.close()
